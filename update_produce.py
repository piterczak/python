#! python3
# update_produce.py - uaktualnia ceny poszczególnych produktów w arkuszu google

import openpyxl


wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# wpisywanie jakie produkty maja zostac uaktualnione i na jakie ceny
Price_updates = {'Garlic':3.07, 'Celery':1.19, 'Lemon':1.27}

for rowNum in range(2, sheet.max_row): #pomijanie 1 rzędu
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in Price_updates:
        sheet.cell(row=rowNum, column = 2).value = Price_updates[produceName]

wb.save('updated_produce_sales.xlsx')
print('Workbook updated and saved to a new file!')